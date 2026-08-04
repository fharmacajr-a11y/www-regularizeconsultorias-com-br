#!/usr/bin/env python3
"""Importa a lista oficial de vagas do Farmacia Popular para JSON."""

from __future__ import annotations

import argparse
import gzip
import hashlib
import json
import os
import re
import sys
import tempfile
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable
from urllib.request import Request, urlopen

try:
    import openpyxl
except ImportError as exc:  # pragma: no cover - depende do ambiente de execucao
    raise SystemExit("Dependencia ausente: instale openpyxl para ler o XLSX.") from exc


IMPORTADOR_VERSAO = "1.1.0"
ABA_ESPERADA = "Planilha2"
COLUNAS_OBRIGATORIAS = (
    "Região",
    "UF",
    "Cód. IBGE",
    "Município",
    "Vagas totais",
    "Vagas preenchidas",
    "Vagas disponíveis",
)
URL_IBGE = "https://servicodados.ibge.gov.br/api/v1/localidades/municipios"
URL_OFICIAL_VISUALIZACAO = (
    "https://www.gov.br/saude/pt-br/composicao/sectics/"
    "farmacia-popular/credenciamento"
)
URL_OFICIAL_DOWNLOAD = (
    "https://www.gov.br/saude/pt-br/composicao/sectics/"
    "farmacia-popular/credenciamento/documentacao/"
    "anexo-i-lista-de-municipios-atualizada-em-28-07-2026.xlsx"
)


class ImportacaoError(ValueError):
    """Erro de entrada ou de consistencia que impede a geracao da base."""


@dataclass(frozen=True)
class TotaisEsperados:
    registros: int
    ufs: int
    vagas_totais: int
    vagas_preenchidas: int
    vagas_disponiveis: int


TOTAIS_BASE_2026_07_28 = TotaisEsperados(
    registros=1082,
    ufs=26,
    vagas_totais=1644,
    vagas_preenchidas=0,
    vagas_disponiveis=1644,
)

REGIOES_EXIBICAO = {
    "NORTE": "Norte",
    "NORDESTE": "Nordeste",
    "CENTRO-OESTE": "Centro-Oeste",
    "SUDESTE": "Sudeste",
    "SUL": "Sul",
}

AVISOS_BASE = [
    "O PDF publicado nao foi usado como fonte do JSON.",
    "O XLSX da fonte usa codigos IBGE de seis digitos; o setimo digito foi obtido por correspondencia de codigo na referencia IBGE.",
]

VALIDACAO_PDF = {
    "arquivo": "PDF publicado na pagina oficial de credenciamento",
    "nomes_a_revisar": [
        "Passa-Vinte -> Passa Vinte",
        "São Luiz -> São Luiz do Anauá",
        "Fortaleza do Tabocão -> Tabocão",
    ],
}


def sha256_arquivo(caminho: Path) -> str:
    digest = hashlib.sha256()
    with caminho.open("rb") as arquivo:
        for bloco in iter(lambda: arquivo.read(1024 * 1024), b""):
            digest.update(bloco)
    return digest.hexdigest()


def _texto_nao_vazio(valor: Any, campo: str, linha: int) -> str:
    if valor is None or not str(valor).strip():
        raise ImportacaoError(f"registro incompleto na linha {linha}: {campo} vazio")
    return str(valor)


def normalizar_nome_para_comparacao(nome: str) -> str:
    """Normaliza nomes somente para comparar a grafia do MS com a do IBGE."""
    apostrofos_tipograficos = "\u2018\u2019\u201b\u2032\u02bc"
    texto = str(nome).casefold().translate(str.maketrans({caractere: "'" for caractere in apostrofos_tipograficos}))
    texto = "".join(" " if unicodedata.category(caractere) == "Pd" else caractere for caractere in texto)
    texto = "".join(
        caractere
        for caractere in unicodedata.normalize("NFD", texto)
        if not unicodedata.combining(caractere)
    )
    return " ".join(texto.split())


def comparar_nomes_municipio(nome_fonte_ms: str, municipio_exibicao: str) -> bool:
    """Informa se os nomes são equivalentes após a normalização de comparação."""
    return normalizar_nome_para_comparacao(nome_fonte_ms) == normalizar_nome_para_comparacao(
        municipio_exibicao
    )


def resumir_comparacao_nominal(registros: Iterable[dict[str, Any]]) -> dict[str, Any]:
    """Calcula métricas e lista apenas as divergências nominais relevantes."""
    registros = list(registros)
    divergencias_relevantes = [
        {
            "codigo_ibge": registro["codigo_ibge"],
            "uf": registro["uf"],
            "municipio_fonte_ms": registro["municipio_fonte_ms"],
            "municipio_exibicao": registro["municipio_exibicao"],
        }
        for registro in registros
        if not comparar_nomes_municipio(
            registro["municipio_fonte_ms"], registro["municipio_exibicao"]
        )
    ]
    return {
        "quantidade_nomes_com_diferenca_literal": sum(
            registro["municipio_fonte_ms"] != registro["municipio_exibicao"]
            for registro in registros
        ),
        "quantidade_nomes_equivalentes_apos_normalizacao": len(registros) - len(divergencias_relevantes),
        "quantidade_divergencias_nominais_relevantes": len(divergencias_relevantes),
        "divergencias_nominais_relevantes": divergencias_relevantes,
    }


def _codigo_numerico(valor: Any, campo: str, linha: int) -> str:
    if isinstance(valor, bool) or valor is None:
        raise ImportacaoError(f"codigo IBGE invalido na linha {linha}: {valor!r}")
    if isinstance(valor, float):
        if not valor.is_integer():
            raise ImportacaoError(f"codigo IBGE invalido na linha {linha}: {valor!r}")
        valor = int(valor)
    texto = str(valor).strip()
    if texto.endswith(".0") and texto[:-2].isdigit():
        texto = texto[:-2]
    if not re.fullmatch(r"\d+", texto) or len(texto) not in (6, 7):
        raise ImportacaoError(
            f"codigo IBGE invalido na linha {linha}: {valor!r}; esperado codigo de 6 ou 7 digitos"
        )
    return texto.zfill(7) if len(texto) == 7 else texto


def _inteiro_nao_negativo(valor: Any, campo: str, linha: int) -> int:
    if isinstance(valor, bool) or valor is None or (isinstance(valor, str) and not valor.strip()):
        raise ImportacaoError(f"{campo} vazio ou invalido na linha {linha}")
    if isinstance(valor, int):
        numero = valor
    elif isinstance(valor, float) and valor.is_integer():
        numero = int(valor)
    elif isinstance(valor, str) and re.fullmatch(r"[+-]?\d+", valor.strip()):
        numero = int(valor.strip())
    else:
        raise ImportacaoError(f"{campo} nao inteiro na linha {linha}: {valor!r}")
    if numero < 0:
        raise ImportacaoError(f"{campo} negativo na linha {linha}: {numero}")
    return numero


def _extrair_uf(item: dict[str, Any]) -> str | None:
    def procurar(valor: Any) -> str | None:
        if isinstance(valor, dict):
            uf = valor.get("uf", valor.get("UF"))
            if isinstance(uf, dict):
                uf = uf.get("sigla")
            if uf:
                return str(uf).strip().upper()
            for filho in valor.values():
                resultado = procurar(filho)
                if resultado:
                    return resultado
        elif isinstance(valor, list):
            for filho in valor:
                resultado = procurar(filho)
                if resultado:
                    return resultado
        return None

    resultado = procurar(item)
    if resultado:
        return resultado
    return None


def _carregar_referencia(payload: Any) -> dict[str, dict[str, str]]:
    if isinstance(payload, dict):
        for chave in ("municipios", "data", "results"):
            if isinstance(payload.get(chave), list):
                payload = payload[chave]
                break
    if not isinstance(payload, list):
        raise ImportacaoError("referencia IBGE deve ser uma lista de municipios")

    referencia: dict[str, dict[str, str]] = {}
    for item in payload:
        if not isinstance(item, dict):
            raise ImportacaoError("item invalido na referencia IBGE")
        codigo = _codigo_numerico(item.get("id", item.get("codigo_ibge")), "id IBGE", 0)
        if len(codigo) != 7:
            raise ImportacaoError(f"id IBGE invalido na referencia: {codigo}")
        nome = item.get("nome", item.get("municipio_exibicao"))
        uf = _extrair_uf(item)
        if nome is None or not str(nome).strip() or not uf:
            raise ImportacaoError(f"referencia IBGE incompleta para o codigo {codigo}")
        if codigo in referencia:
            raise ImportacaoError(f"codigo IBGE duplicado na referencia: {codigo}")
        referencia[codigo] = {"nome": str(nome), "uf": uf}
    if not referencia:
        raise ImportacaoError("referencia IBGE vazia")
    return referencia


def carregar_referencia_json(caminho: Path) -> dict[str, dict[str, str]]:
    try:
        payload = json.loads(caminho.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        raise ImportacaoError(f"nao foi possivel ler a referencia IBGE: {caminho}") from exc
    return _carregar_referencia(payload)


def consultar_referencia_ibge(url: str = URL_IBGE) -> dict[str, dict[str, str]]:
    requisicao = Request(url, headers={"User-Agent": "regularize-consultoria-importador/1.0"})
    try:
        with urlopen(requisicao, timeout=30) as resposta:
            conteudo = resposta.read()
            if resposta.headers.get("Content-Encoding", "").lower() == "gzip" or conteudo.startswith(b"\x1f\x8b"):
                conteudo = gzip.decompress(conteudo)
            payload = json.loads(conteudo.decode("utf-8"))
    except Exception as exc:  # pragma: no cover - rede e servidor externo
        raise ImportacaoError(f"falha ao consultar a referencia IBGE: {url}") from exc
    return _carregar_referencia(payload)


def _resolver_codigo(codigo_fonte: str, referencia: dict[str, dict[str, str]], linha: int) -> str:
    if len(codigo_fonte) == 7:
        if codigo_fonte not in referencia:
            raise ImportacaoError(
                f"codigo IBGE nao encontrado na referencia na linha {linha}: {codigo_fonte}"
            )
        return codigo_fonte
    candidatos = [codigo for codigo in referencia if codigo.startswith(codigo_fonte)]
    if len(candidatos) != 1:
        if not candidatos:
            raise ImportacaoError(
                f"codigo IBGE nao encontrado na referencia na linha {linha}: {codigo_fonte}"
            )
        raise ImportacaoError(f"codigo IBGE ambiguo na referencia na linha {linha}: {codigo_fonte}")
    return candidatos[0]


def _linhas_planilha(caminho: Path) -> Iterable[tuple[int, tuple[Any, ...]]]:
    try:
        workbook = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    except Exception as exc:
        raise ImportacaoError(f"nao foi possivel ler o XLSX: {caminho}") from exc
    if ABA_ESPERADA not in workbook.sheetnames:
        raise ImportacaoError(
            f"aba esperada ausente: {ABA_ESPERADA}; abas encontradas: {workbook.sheetnames}"
        )
    worksheet = workbook[ABA_ESPERADA]
    cabecalho = None
    posicoes: dict[str, int] = {}
    for numero_linha, linha in enumerate(worksheet.iter_rows(max_row=20, values_only=True), 1):
        valores = [str(valor).strip() if valor is not None else "" for valor in linha]
        if set(COLUNAS_OBRIGATORIAS).issubset(valores):
            cabecalho = numero_linha
            posicoes = {coluna: valores.index(coluna) for coluna in COLUNAS_OBRIGATORIAS}
            break
    if cabecalho is None:
        raise ImportacaoError(
            "colunas obrigatorias ausentes ou renomeadas; esperado: "
            + ", ".join(COLUNAS_OBRIGATORIAS)
        )
    for numero_linha, linha in enumerate(
        worksheet.iter_rows(min_row=cabecalho + 1, values_only=True), cabecalho + 1
    ):
        if not any(valor is not None and str(valor).strip() for valor in linha):
            continue
        yield numero_linha, tuple(linha[posicoes[coluna]] if posicoes[coluna] < len(linha) else None for coluna in COLUNAS_OBRIGATORIAS)


def importar_registros(
    caminho_xlsx: Path,
    referencia: dict[str, dict[str, str]],
    esperados: TotaisEsperados = TOTAIS_BASE_2026_07_28,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    prefixos: dict[str, list[str]] = {}
    for codigo in referencia:
        prefixos.setdefault(codigo[:6], []).append(codigo)

    registros: list[dict[str, Any]] = []
    codigos_vistos: set[str] = set()
    for numero_linha, linha in _linhas_planilha(caminho_xlsx):
        regiao_fonte, uf_fonte, codigo_fonte, nome_fonte, total, preenchidas, disponiveis = linha
        regiao = _texto_nao_vazio(regiao_fonte, "Região", numero_linha).strip()
        uf = _texto_nao_vazio(uf_fonte, "UF", numero_linha).strip().upper()
        nome = _texto_nao_vazio(nome_fonte, "Município", numero_linha)
        codigo_bruto = _codigo_numerico(codigo_fonte, "Cód. IBGE", numero_linha)
        if len(codigo_bruto) == 6:
            candidatos = prefixos.get(codigo_bruto, [])
            if len(candidatos) != 1:
                raise ImportacaoError(
                    f"codigo IBGE nao encontrado ou ambiguo na referencia na linha {numero_linha}: {codigo_bruto}"
                )
            codigo = candidatos[0]
        else:
            codigo = _resolver_codigo(codigo_bruto, referencia, numero_linha)
        if codigo in codigos_vistos:
            raise ImportacaoError(f"codigo IBGE duplicado na planilha na linha {numero_linha}: {codigo}")
        codigos_vistos.add(codigo)
        municipio = referencia[codigo]
        if uf != municipio["uf"]:
            raise ImportacaoError(
                f"UF incompatível na linha {numero_linha}: planilha={uf}, IBGE={municipio['uf']} para {codigo}"
            )
        vagas_totais = _inteiro_nao_negativo(total, "Vagas totais", numero_linha)
        vagas_preenchidas = _inteiro_nao_negativo(preenchidas, "Vagas preenchidas", numero_linha)
        vagas_disponiveis = _inteiro_nao_negativo(disponiveis, "Vagas disponíveis", numero_linha)
        if vagas_totais != vagas_preenchidas + vagas_disponiveis:
            raise ImportacaoError(
                f"total de vagas inconsistente na linha {numero_linha}: "
                f"{vagas_totais} != {vagas_preenchidas} + {vagas_disponiveis}"
            )
        registros.append(
            {
                "codigo_ibge": codigo,
                "regiao": REGIOES_EXIBICAO.get(regiao.upper(), regiao),
                "uf": uf,
                "municipio_fonte_ms": nome,
                "municipio_exibicao": municipio["nome"],
                "vagas_totais": vagas_totais,
                "vagas_preenchidas": vagas_preenchidas,
                "vagas_disponiveis": vagas_disponiveis,
            }
        )

    totais = {
        "registros": len(registros),
        "ufs": len({registro["uf"] for registro in registros}),
        "vagas_totais": sum(registro["vagas_totais"] for registro in registros),
        "vagas_preenchidas": sum(registro["vagas_preenchidas"] for registro in registros),
        "vagas_disponiveis": sum(registro["vagas_disponiveis"] for registro in registros),
    }
    esperados_dict = {
        "registros": esperados.registros,
        "ufs": esperados.ufs,
        "vagas_totais": esperados.vagas_totais,
        "vagas_preenchidas": esperados.vagas_preenchidas,
        "vagas_disponiveis": esperados.vagas_disponiveis,
    }
    for campo, esperado in esperados_dict.items():
        if totais[campo] != esperado:
            raise ImportacaoError(
                f"{campo} inesperado: recebido={totais[campo]}, esperado={esperado}"
            )
    registros.sort(key=lambda registro: (registro["uf"], registro["municipio_exibicao"], registro["codigo_ibge"]))
    return registros, resumir_comparacao_nominal(registros)


def escrever_json(caminho: Path, dados: Any) -> None:
    conteudo = json.dumps(dados, ensure_ascii=False, indent=2) + "\n"
    caminho.parent.mkdir(parents=True, exist_ok=True)
    temporario = None
    try:
        with tempfile.NamedTemporaryFile("w", encoding="utf-8", dir=caminho.parent, delete=False) as arquivo:
            temporario = Path(arquivo.name)
            arquivo.write(conteudo)
        os.replace(temporario, caminho)
    finally:
        if temporario and temporario.exists():
            temporario.unlink()


def criar_metadados(
    caminho_xlsx: Path,
    caminho_referencia: Path | None,
    registros: list[dict[str, Any]],
    comparacao_nominal: dict[str, Any],
    data_importacao: str,
    origem_ibge: str,
    versao_referencia_ibge: str,
) -> dict[str, Any]:
    return {
        "orgao_origem": "Ministério da Saúde",
        "titulo_oficial": "Anexo I - Lista de municípios - atualizada em 28/07/2026",
        "data_oficial": "2026-07-28",
        "url_oficial_visualizacao": URL_OFICIAL_VISUALIZACAO,
        "url_oficial_download": URL_OFICIAL_DOWNLOAD,
        "data_importacao": data_importacao,
        "sha256_xlsx": sha256_arquivo(caminho_xlsx),
        "origem_nomes_ibge": origem_ibge,
        "data_ou_versao_referencia_ibge": versao_referencia_ibge,
        "sha256_referencia_ibge": sha256_arquivo(caminho_referencia) if caminho_referencia else None,
        "quantidade_registros": len(registros),
        "quantidade_ufs": len({registro["uf"] for registro in registros}),
        "totais_vagas": {
            "vagas_totais": sum(registro["vagas_totais"] for registro in registros),
            "vagas_preenchidas": sum(registro["vagas_preenchidas"] for registro in registros),
            "vagas_disponiveis": sum(registro["vagas_disponiveis"] for registro in registros),
        },
        "versao_importador": IMPORTADOR_VERSAO,
        "avisos": AVISOS_BASE,
        "validacao_pdf": VALIDACAO_PDF,
        **comparacao_nominal,
    }


def criar_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--xlsx", type=Path, required=True, help="caminho local do XLSX do Ministério da Saúde")
    referencia = parser.add_mutually_exclusive_group(required=True)
    referencia.add_argument("--ibge-json", type=Path, help="JSON local previamente obtido da API do IBGE")
    referencia.add_argument("--consultar-ibge", action="store_true", help="consulta a API do IBGE explicitamente")
    parser.add_argument("--ibge-url", default=URL_IBGE, help=argparse.SUPPRESS)
    parser.add_argument("--ibge-versao", required=True)
    parser.add_argument("--saida", type=Path, required=True, help="caminho do JSON de saída")
    parser.add_argument("--metadados", type=Path, help="caminho do metadados.json")
    parser.add_argument("--data-importacao", required=True)
    return parser


def main(argv: list[str] | None = None) -> int:
    argumentos = sys.argv[1:] if argv is None else argv
    parser = criar_parser()
    if not argumentos:
        parser.print_help()
        return 2
    try:
        opcoes = parser.parse_args(argumentos)
        if opcoes.consultar_ibge:
            referencia = consultar_referencia_ibge(opcoes.ibge_url)
            origem_ibge = f"API oficial do IBGE: {opcoes.ibge_url}"
            caminho_referencia = None
        else:
            referencia = carregar_referencia_json(opcoes.ibge_json)
            origem_ibge = f"JSON local previamente obtido da API oficial do IBGE: {opcoes.ibge_json.name}"
            caminho_referencia = opcoes.ibge_json
        registros, comparacao_nominal = importar_registros(opcoes.xlsx, referencia)
        metadados = criar_metadados(
            opcoes.xlsx,
            caminho_referencia,
            registros,
            comparacao_nominal,
            opcoes.data_importacao,
            origem_ibge,
            opcoes.ibge_versao,
        )
        escrever_json(opcoes.saida, registros)
        escrever_json(opcoes.metadados or opcoes.saida.parent / "metadados.json", metadados)
    except (ImportacaoError, OSError, ValueError) as exc:
        print(f"Erro: {exc}", file=sys.stderr)
        return 1
    print(f"Gerados {len(registros)} registros em {opcoes.saida}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
